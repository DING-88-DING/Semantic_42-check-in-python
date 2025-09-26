import requests
from web3 import Web3
from eth_account.messages import encode_defunct
from openpyxl import load_workbook, Workbook
import logging
from datetime import datetime, timezone
import uuid

# --- 配置日志记录 ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- 全局配置 ---
PRIVY_HEADERS = {
    "accept": "application/json",
    "accept-language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,zh-TW;q=0.6",
    "cache-control": "no-cache",
    "content-type": "application/json",
    "origin": "https://42.semanticlayer.io",
    "pragma": "no-cache",
    "priority": "u=1, i",
    "privy-app-id": "cmfp54at3009nky0bxeg3kqy5",
    "privy-client": "react-auth:2.18.1",
    "referer": "https://42.semanticlayer.io/",
    "sec-ch-ua": '"Not;A=Brand";v="99", "Google Chrome";v="139", "Chromium";v="139"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "cross-site",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36"
}
COPIUM_HEADERS = {
    "accept": "*/*",
    "accept-language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,zh-TW;q=0.6",
    "cache-control": "no-cache",
    "content-type": "application/json",
    "origin": "https://42.semanticlayer.io",
    "pragma": "no-cache",
    "priority": "u=1, i",
    "referer": "https://42.semanticlayer.io/",
    "sec-ch-ua": '"Not;A=Brand";v="99", "Google Chrome";v="139", "Chromium";v="139"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-site",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36"
}
EXCEL_FILE = "wallets.xlsx"

def process_wallet(private_key: str, proxy_ip: str = None):
    try:
        if not private_key.startswith('0x'):
            private_key = '0x' + private_key
        account = Web3().eth.account.from_key(private_key)
        address = account.address
        logging.info(f"钱包地址: {address}")
    except Exception as e:
        logging.error(f"无效的私钥: {private_key[:10]}... - 错误: {e}")
        return

    proxies = None
    if proxy_ip:
        proxies = {"http": f"http://{proxy_ip}", "https": f"http://{proxy_ip}"}
        logging.info(f"使用代理: {proxy_ip}")
    else:
        logging.info("不使用代理")

    # --- 动态生成本次会话的ID ---
    correlation_id = str(uuid.uuid4())
    dynamic_privy_headers = PRIVY_HEADERS.copy()
    dynamic_privy_headers['privy-ca-id'] = correlation_id
    logging.info(f"为本次会话生成新的 privy-ca-id: {correlation_id}")

    privy_session = requests.Session()
    privy_session.proxies = proxies
    privy_session.headers.update(dynamic_privy_headers)

    try:
        logging.info("步骤 1/5: 正在获取Nonce...")
        init_payload = {"address": address}
        init_response = privy_session.post("https://auth.privy.io/api/v1/siwe/init", json=init_payload, timeout=30)
        init_response.raise_for_status()
        nonce = init_response.json()['nonce']
        logging.info(f"成功获取Nonce: {nonce[:15]}...")

        logging.info("步骤 2/5: 正在构建并签名消息...")
        issued_at = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
        
        siwe_message = (
            f"42.semanticlayer.io wants you to sign in with your Ethereum account:\n"
            f"{address}\n"
            f"\n"
            f"By signing, you are proving you own this wallet and logging in. This does not initiate a transaction or cost any fees.\n"
            f"\n"
            f"URI: https://42.semanticlayer.io\n"
            f"Version: 1\n"
            f"Chain ID: 8453\n"
            f"Nonce: {nonce}\n"
            f"Issued At: {issued_at}\n"
            f"Resources:\n"
            f"- https://privy.io"
        )
        
        message_hash = encode_defunct(text=siwe_message)
        signed_message = account.sign_message(message_hash)
        signature = signed_message.signature.hex()
        logging.info("消息签名成功")

        logging.info("步骤 3/5: 正在认证并获取Token...")
        auth_payload = {
            "message": siwe_message,
            "signature": signature,
            "chainId": "eip155:8453",
            "walletClientType": "okx_wallet",
            "connectorType": "injected",
            "mode": "login-or-sign-up"
        }
        auth_response = privy_session.post("https://auth.privy.io/api/v1/siwe/authenticate", json=auth_payload, timeout=30)
        auth_response.raise_for_status()
        privy_token = auth_response.json()['token']
        logging.info("成功获取Privy Token")

        logging.info("步骤 4/5: 正在初始化用户...")
        init_user_payload = {"data": {"address": address, "privyToken": privy_token}}
        init_user_response = requests.post("https://copium-api.semanticlayer.io/initializeuser", headers=COPIUM_HEADERS, json=init_user_payload, proxies=proxies, timeout=30)
        logging.info(f"初始化用户响应: {init_user_response.status_code} - {init_user_response.text}")
        init_user_response.raise_for_status()
        logging.info("用户初始化成功或已存在")

        logging.info("步骤 5/5: 正在执行每日签到...")
        claim_payload = {"data": {"address": address, "privyToken": privy_token}}
        claim_response = requests.post("https://copium-api.semanticlayer.io/claimdailyxp", headers=COPIUM_HEADERS, json=claim_payload, proxies=proxies, timeout=30)
        
        if claim_response.status_code == 200:
            logging.info(f"✅ 签到成功! 地址: {address} - 响应: {claim_response.json()}")
        else:
            logging.error(f"❌ 签到失败! 地址: {address} - 状态码: {claim_response.status_code} - 响应: {claim_response.text}")

    except requests.exceptions.RequestException as e:
        logging.error(f"❌ 网络请求失败! 地址: {address} - 错误: {e}")
    except Exception as e:
        logging.error(f"❌ 发生未知错误! 地址: {address} - 错误: {e}")

def main():
    try:
        workbook = load_workbook(filename=EXCEL_FILE)
        sheet = workbook.active
    except FileNotFoundError:
        logging.warning(f"未找到 '{EXCEL_FILE}' 文件。将为您创建一个新的模板。")
        wb = Workbook()
        ws = wb.active
        ws.title = "Wallets"
        ws.append(["PrivateKey", "ProxyIP"])
        ws.append(["0xaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa", "127.0.0.1:7890"])
        ws.append(["0xbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb", "user:pass@host:port"])
        wb.save(EXCEL_FILE)
        logging.info(f"模板文件 '{EXCEL_FILE}' 已创建。请填入您的私钥和可选的代理信息后重新运行。")
        return

    for row in sheet.iter_rows(min_row=2, values_only=True):
        private_key = row[0] if len(row) > 0 else None
        proxy_ip = row[1] if len(row) > 1 else None

        if not private_key:
            logging.warning(f"跳过缺少私钥的数据行: {row}")
            continue

        process_wallet(str(private_key), str(proxy_ip) if proxy_ip else None)
        logging.info("-" * 70)

if __name__ == "__main__":
    main()