import requests
from web3 import Web3
from eth_account.messages import encode_defunct
from openpyxl import load_workbook, Workbook
import logging
from datetime import datetime, timezone
import uuid

# --- 配置日志记录 ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- 全局配置 ---
PRIVY_HEADERS = {
    "accept": "application/json",
    "accept-language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,zh-TW;q=0.6",
    "cache-control": "no-cache",
    "content-type": "application/json",
    "origin": "https://42.semanticlayer.io",
    "pragma": "no-cache",
    "priority": "u=1, i",
    "privy-app-id": "cmfp54at3009nky0bxeg3kqy5",
    "privy-client": "react-auth:2.18.1",
    "referer": "https://42.semanticlayer.io/",
    "sec-fetch-site": "cross-site",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36"
}

COPIUM_HEADERS = {
    "accept": "*/*",
    "content-type": "application/json",
    "origin": "https://42.semanticlayer.io",
    "referer": "https://42.semanticlayer.io/",
    "sec-fetch-site": "same-site",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36"
}

EXCEL_FILE = "wallets.xlsx"


def process_wallet(private_key: str, proxy_ip: str = None):
    # 私钥处理
    try:
        if not private_key.startswith("0x"):
            private_key = "0x" + private_key

        account = Web3().eth.account.from_key(private_key)
        address = account.address
        logging.info(f"钱包地址: {address}")

    except Exception as e:
        logging.error(f"无效的私钥: {private_key[:10]}... 错误: {e}")
        return

    # 代理
    proxies = None
    if proxy_ip:
        proxies = {"http": f"http://{proxy_ip}", "https": f"http://{proxy_ip}"}
        logging.info(f"使用代理: {proxy_ip}")
    else:
        logging.info("不使用代理")

    # 动态 privy-ca-id
    correlation_id = str(uuid.uuid4())
    dynamic_headers = PRIVY_HEADERS.copy()
    dynamic_headers["privy-ca-id"] = correlation_id
    logging.info(f"生成 privy-ca-id: {correlation_id}")

    session = requests.Session()
    session.headers.update(dynamic_headers)
    session.proxies = proxies

    try:
        # ------------------------- Step 1 获取 nonce -------------------------
        logging.info("步骤 1/5: 获取 Nonce...")
        init_resp = session.post(
            "https://auth.privy.io/api/v1/siwe/init",
            json={"address": address},
            timeout=30
        )
        init_resp.raise_for_status()
        nonce = init_resp.json()["nonce"]
        logging.info(f"成功获取 nonce: {nonce}")

        # ------------------------- Step 2 构造 siwe_message（严格匹配 JS） -------------------------
        logging.info("步骤 2/5: 生成并签名消息...")

        # JS 的 issuedAt 是：new Date().toISOString()
        issued_at = datetime.now(timezone.utc).isoformat()

        siwe_message = (
            f"42.semanticlayer.io wants you to sign in with your Ethereum account:\n"
            f"{address}\n"
            f"\n"
            f"By signing, you are proving you own this wallet and logging in. This does not initiate a transaction or cost any fees.\n"
            f"\n"
            f"URI: https://42.semanticlayer.io\n"
            f"Version: 1\n"
            f"Chain ID: 8453\n"
            f"Nonce: {nonce}\n"
            f"Issued At: {issued_at}\n"
            f"Resources:\n"
            f"- https://privy.io"
        )

        # JS 使用 account.sign(message)
        message_hash = encode_defunct(text=siwe_message)
        signature = account.sign_message(message_hash).signature.hex()
        logging.info("消息签名成功")

        # ------------------------- Step 3 authenticate（关键修复点） -------------------------
        logging.info("步骤 3/5: 认证并获取 privy token...")

        auth_payload = {
            "message": siwe_message,
            "signature": signature,
            "chainId": "eip155:8453",
            "walletClientType": "okx_wallet",
            "connectorType": "injected",
            "mode": "login-or-sign-up"
        }

        auth_resp = session.post(
            "https://auth.privy.io/api/v1/siwe/authenticate",
            json=auth_payload,
            timeout=30
        )
        auth_resp.raise_for_status()
        privy_token = auth_resp.json()["token"]
        logging.info("成功获取 privy token!")

        # ------------------------- Step 4 初始化用户 -------------------------
        logging.info("步骤 4/5: 初始化用户...")

        init_user_payload = {"data": {"address": address, "privyToken": privy_token}}

        r = requests.post(
            "https://copium-api.semanticlayer.io/initializeuser",
            json=init_user_payload,
            headers=COPIUM_HEADERS,
            proxies=proxies,
            timeout=30
        )
        logging.info(f"初始化响应: {r.status_code} - {r.text}")
        r.raise_for_status()

        # ------------------------- Step 5 签到 -------------------------
        logging.info("步骤 5/5: 每日签到...")

        claim_payload = {"data": {"address": address, "privyToken": privy_token}}

        r2 = requests.post(
            "https://copium-api.semanticlayer.io/claimdailyxp",
            json=claim_payload,
            headers=COPIUM_HEADERS,
            proxies=proxies,
            timeout=30
        )

        if r2.status_code == 200:
            logging.info(f"✅ 签到成功! 地址 {address} - {r2.text}")
        else:
            logging.error(f"❌ 签到失败 Status {r2.status_code}: {r2.text}")

    except Exception as e:
        logging.error(f"❌ 出错: {e}")


def main():
    try:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
    except FileNotFoundError:
        logging.warning("找不到 wallets.xlsx，生成模板...")
        wb = Workbook()
        ws = wb.active
        ws.append(["PrivateKey", "ProxyIP"])
        ws.append(["0xaaaaaaaaaaaaaaaaaaaa", "127.0.0.1:7890"])
        wb.save(EXCEL_FILE)
        return

    for row in sheet.iter_rows(min_row=2, values_only=True):
        pk = row[0]
        proxy = row[1] if len(row) > 1 else None

        if not pk:
            continue

        process_wallet(str(pk), str(proxy) if proxy else None)


if __name__ == "__main__":
    main()


